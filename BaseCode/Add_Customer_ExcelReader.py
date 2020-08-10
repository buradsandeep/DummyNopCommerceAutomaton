import openpyxl

class Xlutils:

    def __init__(self,path, sheetname):
        self.path = path
        self.sheetname = sheetname
        self.workbook = openpyxl.load_workbook(self.path)
        self.worksheet = self.workbook.get_sheet_by_name(self.sheetname)

    def getExcelrows(self):
        return self.worksheet.max_row

    def readExcel(self, row):
            email = self.worksheet.cell(row, 1).value
            password = self.worksheet.cell(row, 2).value
            firstname = self.worksheet.cell(row, 3).value
            lastname = self.worksheet.cell(row, 4).value
            gender =self.worksheet.cell(row, 5).value
            dob = self.worksheet.cell(row, 6).value
            company_name = self.worksheet.cell(row, 7).value
            tax_exemption = self.worksheet.cell(row, 8).value
            newsletter = self.worksheet.cell(row, 9).value
            manage_vendor = self.worksheet.cell(row, 10).value
            active = self.worksheet.cell(row, 11).value
            admin_comment = self.worksheet.cell(row, 12).value
            return email, password, firstname, lastname, gender, dob, company_name, tax_exemption, newsletter, manage_vendor, active, admin_comment

    def writeExcel(self,row,message):
        self.worksheet.cell(row, 13).value = message

    def closeExcel(self):
        self.workbook.save(self.path)
        self.workbook.close()



